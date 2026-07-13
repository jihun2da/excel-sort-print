# -*- coding: utf-8 -*-
"""streamlit 없이 핵심 로직(xlsx_utils, image_utils)만 독립적으로 검증하는 스크립트."""
import sys
import openpyxl

sys.path.insert(0, ".")
from xlsx_utils import load_sheet_rows, sort_sequential, build_multi_sheet_xlsx, COL
from image_utils import build_pages_for_rows

RAW_PATH = "/sessions/sleepy-festive-hypatia/mnt/uploads/콘.xlsx"
REF_PATH = "/sessions/sleepy-festive-hypatia/mnt/uploads/콘-5fa346fa.xlsx"

with open(RAW_PATH, "rb") as f:
    raw_bytes = f.read()

# ---- 1. 대량분류 정렬 vs 사용자 로컬 정렬 결과 비교 ----
loaded = load_sheet_rows(raw_bytes, "Sheet1")
print("Sheet1 데이터 행수:", len(loaded.rows))

bulk_sorted = sort_sequential(loaded.rows, ["A", "I", "J", "K"])

ref_wb = openpyxl.load_workbook(REF_PATH, data_only=True)
ref_ws = ref_wb["Sheet1"]
ref_rows = [r for r in ref_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
print("참조 파일 행수:", len(ref_rows))

mismatches = 0
for a, b in zip(bulk_sorted, ref_rows):
    ac = a["cells"]
    if ac[0] != b[0] or ac[4] != b[4] or ac[8] != b[8] or ac[9] != b[9] or ac[10] != b[10]:
        mismatches += 1
print("대량분류 정렬 불일치 건수 (0이어야 정상):", mismatches)

# ---- 2. 대량분류 서식보존 다운로드 ----
sheet_groups = [{"name": "Sheet1", "row_nums": [r["row_num"] for r in bulk_sorted]}]
out_bytes = build_multi_sheet_xlsx(raw_bytes, "Sheet1", sheet_groups)
with open("test_bulk_out.xlsx", "wb") as f:
    f.write(out_bytes)
print("대량분류 결과 저장 완료, 크기:", len(out_bytes))

check_wb = openpyxl.load_workbook("test_bulk_out.xlsx", data_only=True)
print("결과 시트 목록:", check_wb.sheetnames)
ws_check = check_wb["Sheet1"]
print("결과 시트 차원:", ws_check.dimensions)
r2 = ws_check[2]
print("첫 데이터행 A,E,I,J,K:", r2[0].value, r2[4].value, r2[8].value, r2[9].value, r2[10].value, "fill=", r2[0].fill.fgColor.rgb)

# ---- 3. 소량분류 (F,G,A,K + H열 그룹) : Sheet2로 다중그룹 검증 ----
loaded2 = load_sheet_rows(raw_bytes, "Sheet2")
print("\nSheet2 데이터 행수:", len(loaded2.rows))
small_sorted = sort_sequential(loaded2.rows, ["F", "G", "A", "K"])
threshold = 50
group_map = {}
for row in small_sorted:
    hv = row["cells"][COL["H"]]
    key = "" if hv is None else str(hv)
    group_map.setdefault(key, []).append(row)
large_groups = [(k, v) for k, v in group_map.items() if len(v) >= threshold]
large_keys = {k for k, v in large_groups}
main_rows = [row for row in small_sorted if ("" if row["cells"][COL["H"]] is None else str(row["cells"][COL["H"]])) not in large_keys]
print("largeGroups:", [(k, len(v)) for k, v in large_groups])
print("mainRows:", len(main_rows))

sheets2 = []
if main_rows:
    sheets2.append({"name": "Sheet2_기본", "rows": main_rows})
for k, v in large_groups:
    sheets2.append({"name": k or "(빈값)", "rows": v})

total_check = sum(len(s["rows"]) for s in sheets2)
print("총 행수 보존 체크:", total_check, "== 2443 ?", total_check == 2443)

sheet_groups2 = [{"name": s["name"], "row_nums": [r["row_num"] for r in s["rows"]]} for s in sheets2]
out_bytes2 = build_multi_sheet_xlsx(raw_bytes, "Sheet2", sheet_groups2)
with open("test_small_out.xlsx", "wb") as f:
    f.write(out_bytes2)
print("소량분류 결과 저장 완료, 크기:", len(out_bytes2))

check_wb2 = openpyxl.load_workbook("test_small_out.xlsx", data_only=True)
print("결과 시트 목록:", check_wb2.sheetnames)
for name in check_wb2.sheetnames:
    print("  ", name, check_wb2[name].dimensions)

# ---- 4. 이미지 렌더링 (신규 사양: A열 11pt, 연한 구분선, F/G 자동축소) ----
divider_style = {"color": "#d4d4d4", "width": 1.5}
pages = build_pages_for_rows(bulk_sorted, divider_style)
print("\n대량분류 총 페이지수:", len(pages))
pages[0]["image"].save("test_bulk_page1.png")

# 긴 F/G 값 있는 그룹 페이지도 렌더링 (직배 그룹)
directbae = next((s for s in sheets2 if s["name"] == "직배"), None)
if directbae:
    pages_d = build_pages_for_rows(directbae["rows"], divider_style)
    pages_d[0]["image"].save("test_small_directbae_page1.png")
    print("직배 그룹 페이지수:", len(pages_d))

print("\n=== ALL DONE ===")
